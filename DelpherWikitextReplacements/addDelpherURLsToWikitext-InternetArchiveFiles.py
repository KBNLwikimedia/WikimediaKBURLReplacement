"""
Commons Delpher Source Updater - Adds 2 Delpher URLs to the Wikitext of Internet Archive files
in Category:Media from Delpher
==============================

Batch-updates Wikimedia Commons File pages using an Excel-driven workflow.
For each row in the Excel (Title, Pageid, URL), the script:
  1) Builds a record from the numeric Pageid (strict validation).
  2) Optionally slices the workload via HEAD or RANGE.
  3) Logs into Commons and fetches the page wikitext by pageid.
  4) Searches for EXACTLY ONE “old” Internet Archive source block; if none is found,
     the file is skipped (no change). If multiple are found, the run aborts.
  5) Replaces the old block with a richer block that includes:
       - Internet Archive (template link + direct PDF)
       - Delpher (website + direct PDF via a strict URN: PREFIX:NUMBER)
  6) Saves the edit using a CSRF token, `assert=user`, and `maxlag=5`.
  7) Optionally opens the updated page in a browser tab.
  8) Writes a per-file status back to the Excel sheet.

-----------------------------
Pattern replacement examples

OLD block examples (either form):

|source =
:{{Internet Archive link|ddd_010124175_mpeg21}}
:https://archive.org/download/ddd_010124175_mpeg21/ddd_010124175_mpeg21.pdf

|source =
:{{Internet Archive link|MMKB08_000088804_mpeg21}}
:https://archive.org/download/MMKB08_000088804_mpeg21/MMKB08_000088804_mpeg21.pdf

New block inserted:

|source =
Internet Archive
* Website: {{Internet Archive link|<IA_ID>}}
* Direct download: <IA_PDF_URL>
Delpher
* Website: https://resolver.kb.nl/resolve?urn=<URN_PREFIX>:<NUMBER>
* Direct download: https://resolver.kb.nl/resolve?urn=<URN_PREFIX>:<NUMBER>:mpeg21:pdf

---------------------------------------------
Strictness & Safety
-------------------
- Configuration, input validation, and API responses are strict; most anomalies log
  a CRITICAL message and abort the run (SystemExit). Per-file processing records
  continue across errors so statuses can be written back (processed / skipped).
- The Delpher URN is derived strictly from the IA ID shape: PREFIX_NUMBER[_suffix].
  Leading zeros in NUMBER are preserved.

Statuses written to Excel (EXCEL_STATUS_COL)
--------------------------------------------
- "Successfully processed"         — block replaced and page saved
- "Skipped - no old pattern"       — no old source block to replace
- "Skipped - other error"          — fetch/transform/save error for that page

Configuration (env vars)
------------------------
Required:
- WIKIMEDIA_USERNAME, WIKIMEDIA_PASSWORD, for login to API
Optional (sensible defaults exist):
- WIKIMEDIA_USER_AGENT            : User-Agent sent to Commons API
- EXCEL_FILE                      : Path to the Excel (default: MediaFromDelpher-InternetArchiveFiles_16092025.xlsx)
- EXCEL_SHEET                     : Sheet name or index (int). If empty ⇒ first sheet
- EXCEL_TITLE_COL                 : Title column name (default: "Title")
- EXCEL_URL_COL                   : Commons URL column name (default: "URL")
- EXCEL_PAGEID_COL                : PageID column name (default: "PageID")
- EXCEL_STATUS_COL                : Status column name to write (default: "ProcessingStatus")
- EXCEL_RESOLVER_COL              : Resolver URL, like https://resolver.kb.nl/resolve?urn=ddd:000013084 (default: "ResolverURL")
- MAX_FILES                       : Upper bound on records processed
- EDIT_SUMMARY                    : Commons edit summary
- EDIT_SLEEP_SEC                  : Sleep between files (seconds) to throttle politely
- HEAD                            : Process first N records (mutually exclusive with RANGE)
- RANGE                           : "N-M" (1-based inclusive) slice (mutually exclusive with HEAD)
- OPEN_AFTER_SUCCESS              : 1/0 open a browser tab after each successful edit (default: 1)
- OPEN_AFTER_SUCCESS_MAX          : Max number of tabs to open per run (default: 10)

Dependencies
------------
- mwclient, pandas, openpyxl, python-dotenv, tqdm

Notes
-----
- Designed to run directly from PyCharm (no CLI args).
- Browser open uses Special:Redirect by pageid for robust linking:
  https://commons.wikimedia.org/wiki/Special:Redirect/page/{pageid}

Created by Olaf Janssen, Wikimedia coordinator of KB, national library of the Netherlands
with much help from ChatGPT.
Latest update: 18 September 2025
License = CC0, public domain.
"""


import mwclient
import re
import os
import logging
from typing import List, Dict, Optional
from dotenv import load_dotenv
from tqdm import tqdm
import time
import webbrowser
import pandas as pd
from openpyxl import load_workbook


# ------------------------
# Configuration & Logging
# ------------------------
logging.basicConfig(
    format="%(asctime)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

# Load data from .env file
load_dotenv()

# Wikimedia Commons login credentials (read from environment variables)
USERNAME = os.getenv("WIKIMEDIA_USERNAME", "").strip()
PASSWORD = os.getenv("WIKIMEDIA_PASSWORD", "").strip()
USER_AGENT = os.getenv("WIKIMEDIA_USER_AGENT", "KB-DelpherUpdater/1.0 (olaf.janssen@kb.nl) - Olaf Janssen, KB, national library of the Netherlands").strip()

if not USERNAME or not PASSWORD:
    logging.error(
        "Wikimedia credentials are missing. Please set WIKIMEDIA_USERNAME and WIKIMEDIA_PASSWORD in a .env file or as environment variables.")
    raise SystemExit(1)

if not USER_AGENT:
    logging.warning("User-Agent is not set. Some APIs may reject requests without it.")

# Optional slicing for convenient local runs (no CLI needed)
HEAD = int(os.getenv("HEAD", "0"))   # if >0, take first HEAD entries
RANGE = os.getenv("RANGE", "50000-55759")       # e.g. "51-200" (1-based, inclusive)

# Excel input (always used)
EXCEL_FILE = os.getenv("EXCEL_FILE", "MediaFromDelpher-InternetArchiveFiles_16092025.xlsx")
# If EXCEL_SHEET is unset/blank, use first sheet (index 0). Avoid passing None to pandas to prevent dict return.
_env_sheet = os.getenv("EXCEL_SHEET", "").strip()
EXCEL_SHEET = 0 if not _env_sheet else _env_sheet
EXCEL_TITLE_COL = "Title"   # optional: used for logging only
EXCEL_PAGEID_COL = os.getenv("EXCEL_PAGEID_COL", "PageID") # the column that holds the numeric pageid
EXCEL_URL_COL = "URL"       # preferred: use MediaInfo entity URLs ;  URL = https://commons.wikimedia.org/entity/M<PageID>
EXCEL_RESOLVER_COL = os.getenv("EXCEL_RESOLVER_COL", "ResolverURL")


# Which column we write the processing statuses to the Excel file
EXCEL_STATUS_COL = os.getenv("EXCEL_STATUS_COL", "ProcessingStatus")
# Normalized status strings (change if you prefer other labels)
STATUS_PROCESSED = "Successfully processed"
STATUS_SKIPPED_NO_OLD = "Skipped - no old pattern found"
STATUS_ERROR = "Skipped - other error"

# Limits & edit summary
MAX_FILES = int(os.getenv("MAX_FILES", "60000"))
EDIT_SUMMARY = os.getenv("EDIT_SUMMARY", "Improved sourcing: added Delpher website and direct PDF links")
# Throttle API to avoid overload, pause API writing
EDIT_SLEEP_SEC = float(os.getenv("EDIT_SLEEP_SEC", "4.0"))  # seconds between edits / throttle

OPEN_AFTER_SUCCESS = int(os.getenv("OPEN_AFTER_SUCCESS", "0"))  # 1=on, 0=off, do or do not open modified Commons file page in browser
OPEN_AFTER_SUCCESS_MAX = int(os.getenv("OPEN_AFTER_SUCCESS_MAX", "20"))  # cap how many tabs to open

# Checkpoint (flush) to Excel after this many successful edits - default = 50 files
CHECKPOINT_EVERY_SUCCESS = int(os.getenv("CHECKPOINT_EVERY_SUCCESS", "50"))

#=====================================================================
#=====================================================================

""" 
Regex helpers (pattern transform)
Builds a compiled regular expression named OLD_BLOCK_RE that finds the exact “old” three-line |source = block you want to replace.

|source = 
:{{Internet Archive link|KBDDD02_000201168_mpeg21}}
:https://archive.org/download/KBDDD02_000201168_mpeg21/KBDDD02_000201168_mpeg21.pdf

What it captures (P = named capture groups):
- prefix → the exact |source = line (including its newline).
- iaid → the IA identifier, e.g. KBDDD02_000201168_mpeg21.
- pdf → the direct PDF URL, e.g. https://archive.org/download/KBDDD02_000201168_mpeg21/KBDDD02_000201168_mpeg21.pdf.
"""

OLD_BLOCK_RE = re.compile(
    r"(?P<prefix>\|\s*source\s*=\s*\n)"  # the |source = line including newline
    r"\s*:\s*\{\{\s*Internet\s+Archive\s+link\s*\|\s*(?P<iaid>[^}\n\r]+?)\s*\}\}\s*\n"  # {{Internet Archive link|ID}}
    r"\s*:\s*(?P<pdf>https?://archive\.org/download/[\w\-\./]+?\.pdf)\s*\n?", # https://archive.org/download/....pdf
    re.IGNORECASE)

""" 
Regex for the “ID splitter” that pulls the collection prefix and the zero-padded number out of an Internet Archive ID so we can build a Delpher URN.

Named capture groups:
- prefix → the collection prefix, e.g. KBDDD02
- number → the zero-padded number, e.g. 000201168

Examples:
- ddd_010124175_mpeg21 → prefix='ddd', number='010124175'
- MMKB08_000088804_mpeg21 → prefix='MMKB08', number='000088804'
- KBDDD02_000201168_mpeg21 → prefix='KBDDD02', number='000201168'
We then build the URN as "{prefix}:{number}", so leading zeros are kept:
- KBDDD02_000201168_mpeg21 → KBDDD02:000201168 

"""
URN_FROM_IA_RE = re.compile(r"^(?P<prefix>[A-Za-z0-9]+)_(?P<number>\d+)(?:_.*)?$", re.ASCII)

# Extract IA ID from Title like: "...(IA_MMWFA01_000191129_mpeg21).pdf"
# FROM : Arnhemsche_courant_05-12-1852_(IA_MMKB08_000088804_mpeg21).pdf
# TO:  https://resolver.kb.nl/resolve?urn=MMKB08:000088804
IA_ID_FROM_TITLE_RE = re.compile(r"\(IA_(?P<prefix>[A-Za-z0-9]+)_(?P<number>\d+)(?:_[^)]+)?\)")

def derive_delpher_urn(ia_id: str) -> Optional[str]:
    """
    Strictly derive a Delpher URN from an Internet Archive identifier.

    Expected format (must match exactly):
        PREFIX_NUMBER[_suffix]
    where:
        - PREFIX is ASCII letters/digits, e.g. "ddd", "MMKB08", "KBDDD02"
        - NUMBER is one or more digits (leading zeros preserved), e.g. "010124175", "000201168"
        - Optional trailing suffix (e.g. "_mpeg21") is allowed but ignored

    Examples:
        "ddd_010124175_mpeg21"      -> "ddd:010124175"
        "MMKB08_000088804_mpeg21"   -> "MMKB08:000088804"
        "KBDDD02_000201168_mpeg21"  -> "KBDDD02:000201168"

    On failure:
        - Logs a clear error message indicating the invalid IA ID and the expected shape
        - Halts further processing by exiting the program (SystemExit)

    Returns:
        str: The URN as "{prefix}:{number}" if parsing succeeds.

    Raises:
        SystemExit: If the IA ID does not match the strict pattern.
    """
    if not isinstance(ia_id, str) or not ia_id.strip():
        logging.critical("derive_delpher_urn: empty or invalid IA ID: %r", ia_id)
        raise SystemExit(1)

    s = ia_id.strip()
    m = URN_FROM_IA_RE.match(s)  # ^ and $ are in the pattern; this enforces a full match
    if not m:
        logging.critical(
            "Invalid Internet Archive ID for URN derivation: %r. "
            "Expected format: PREFIX_NUMBER[_suffix], e.g. 'KBDDD02_000201168_mpeg21'.",
            ia_id
        )
        raise SystemExit(1)

    prefix = m.group("prefix")
    number = m.group("number")  # preserve leading zeros
    return f"{prefix}:{number}"

def resolver_from_title(title: Optional[str]) -> Optional[str]:
    """
    Return the Delpher resolver URL from a Commons Title, or None if not found.

    Example:
      Title: "Hoornsche_courant_26-03-1861_(IA_MMWFA01_000191129_mpeg21).pdf"
      -> "https://resolver.kb.nl/resolve?urn=MMWFA01:000191129"
    """
    if not isinstance(title, str) or not title.strip():
        return None
    m = IA_ID_FROM_TITLE_RE.search(title)
    if not m:
        return None
    prefix = m.group("prefix")
    number = m.group("number")  # keep leading zeros
    return f"https://resolver.kb.nl/resolve?urn={prefix}:{number}"


def build_new_source_block(ia_id: str, pdf_url: str) -> Optional[str]:

    """
    Strictly build the replacement |source= block from an Internet Archive ID and its direct PDF URL.

    Requirements (no fallback parsing):
      - ia_id must match URN_FROM_IA_RE exactly:
            PREFIX_NUMBER[_suffix]
        where PREFIX = [A-Za-z0-9]+ and NUMBER = \\d+ (leading zeros preserved).
      - pdf_url must be an Archive.org direct download URL ending in ".pdf", e.g.:
            https://archive.org/download/KBDDD02_000201168_mpeg21/KBDDD02_000201168_mpeg21.pdf

    On invalid inputs:
      - Logs a CRITICAL error with the offending value.
      - Halts execution (SystemExit).

    Returns:
      The newline-joined block string (never None unless execution is aborted).
    """
    # Validate basic types/emptiness
    if not isinstance(ia_id, str) or not ia_id.strip():
        logging.critical("build_new_source_block: empty or invalid ia_id: %r", ia_id)
        raise SystemExit(1)
    if not isinstance(pdf_url, str) or not pdf_url.strip():
        logging.critical("build_new_source_block: empty or invalid pdf_url: %r", pdf_url)
        raise SystemExit(1)

    s_pdf = pdf_url.strip()

    # Strict PDF URL check (anchor to full string)
    if not re.match(r'^https?://archive\.org/download/[\w\-\./]+?\.pdf$', s_pdf):
        logging.critical(
            "Invalid PDF URL for Internet Archive direct download: %r. "
            "Expected e.g. 'https://archive.org/download/<path>/<file>.pdf'",
            pdf_url
        )
        raise SystemExit(1)

    # derive_delpher_urn is already strict and will SystemExit on mismatch
    urn = derive_delpher_urn(ia_id)

    ia_link_tpl = f"{{{{Internet Archive link|{ia_id}}}}}"
    delpher_site = f"https://resolver.kb.nl/resolve?urn={urn}"
    delpher_pdf  = f"https://resolver.kb.nl/resolve?urn={urn}:mpeg21:pdf"

    return "\n".join([
        "Internet Archive",
        f"* Website: {ia_link_tpl}",
        f"* Direct download: {s_pdf}",
        "Delpher",
        f"* Website: {delpher_site}",
        f"* Direct download: {delpher_pdf}"
    ])

def transform_wikitext(text: str) -> str:
    """
    Transform exactly one OLD |source= block into the new format.
    - If there are ZERO matches: log and return the original text unchanged (skip this file).
    - If there are MULTIPLE matches: log CRITICAL and abort (ambiguous).
    - If there is EXACTLY ONE match: replace it, **logging the derived URN**, and return the updated text.
    """
    if not isinstance(text, str):
        logging.critical("transform_wikitext: 'text' must be a str, got %r", type(text).__name__)
        raise SystemExit(1)

    matches = list(OLD_BLOCK_RE.finditer(text))
    if len(matches) == 0:
        logging.info("transform_wikitext: no OLD |source= block found; skipping this file.")
        return text
    if len(matches) > 1:
        logging.critical("transform_wikitext: multiple (%d) OLD |source= blocks found; aborting.", len(matches))
        raise SystemExit(1)

    m = matches[0]
    prefix = m.group("prefix")
    iaid   = m.group("iaid").strip()
    pdf    = m.group("pdf").strip()

    # Strict URN derivation (will exit on mismatch) + log it
    urn = derive_delpher_urn(iaid)
    logging.info("transform_wikitext: Derived URN %r from IA ID %r", urn, iaid)

    # Build the new block (strict; will exit on invalid inputs)
    new_block = build_new_source_block(iaid, pdf)

    # Replace only the matched span; keep an extra newline after the new block
    start, end = m.span()
    return text[:start] + prefix + new_block + "\n" + text[end:]


# ------------------------
# Excel helpers
# ------------------------

def apply_slice_records(records: List[Dict]) -> List[Dict]:
    """
    Strictly select a subset of records based on environment variables RANGE or HEAD.

    Rules (no fallback / no silent clamping):
      - Exactly one of RANGE or HEAD may be set. If both are set, abort.
      - RANGE: a 1-based inclusive span in the form "N-M".
          * Validates format strictly.
          * Requires 1 <= N <= M <= len(records).
          * Returns records[N-1:M].
      - HEAD: a positive integer count.
          * Requires 1 <= HEAD <= len(records).
          * Returns records[:HEAD].
      - If neither is set, returns all records unchanged.

    On invalid configuration, logs a CRITICAL message and aborts (SystemExit).
    """
    if not isinstance(records, list):
        logging.critical("apply_slice_records: 'records' must be a list, got %r", type(records).__name__)
        raise SystemExit(1)

    total = len(records)
    if total == 0:
        return records

    has_range = bool(RANGE and str(RANGE).strip())
    has_head  = bool(HEAD and int(HEAD) > 0)

    if has_range and has_head:
        logging.critical("apply_slice_records: both RANGE and HEAD are set; set only one.")
        raise SystemExit(1)

    if has_range:
        m = re.match(r"^\s*(\d+)\s*[-:]\s*(\d+)\s*$", str(RANGE))
        if not m:
            logging.critical("apply_slice_records: RANGE must be in the form N-M (1-based). Got: %r", RANGE)
            raise SystemExit(1)
        start = int(m.group(1))
        end   = int(m.group(2))
        if start < 1 or end < start:
            logging.critical("apply_slice_records: invalid RANGE values: start=%d, end=%d. Must satisfy 1 <= start <= end.", start, end)
            raise SystemExit(1)
        if start > total or end > total:
            logging.critical("apply_slice_records: RANGE out of bounds for %d records (start=%d, end=%d).", total, start, end)
            raise SystemExit(1)
        i, j = start - 1, end
        logging.info("apply_slice_records: using RANGE %d-%d (inclusive) out of %d.", start, end, total)
        return records[i:j]

    if has_head:
        n = int(HEAD)
        if n < 1 or n > total:
            logging.critical("apply_slice_records: HEAD must be between 1 and %d. Got: %d", total, n)
            raise SystemExit(1)
        logging.info("apply_slice_records: using HEAD=%d out of %d.", n, total)
        return records[:n]

    logging.info("apply_slice_records: no RANGE/HEAD set; processing all %d records.", total)
    return records


# ------------------------
# MediaWiki API helpers (edit by pageid)
# ------------------------

def fetch_wikitext_by_pageid(site: mwclient.Site, pageid: int) -> Dict[str, str]:
    """
    Strictly fetch the current wikitext and base timestamp for a single File page by pageid.

    Requirements / Guarantees:
      - 'pageid' must be a positive integer.
      - Uses the MediaWiki API: action=query with formatversion=2 and rvslots=main.
      - Returns a dict with keys:
          { "text": <wikitext str>, "basetimestamp": <ISO8601 str> }
      - If anything is unexpected (missing page, no revisions, no main slot, etc.), the
        function logs a CRITICAL error and halts the program (SystemExit). No silent None.

    Args:
      site: Authenticated mwclient.Site for Commons.
      pageid: Positive integer page ID in namespace 6 (File).

    Returns:
      dict: {"text": str, "basetimestamp": str}

    Raises:
      SystemExit: on invalid input or any API / data validation failure.
    """
    # Validate input
    if not isinstance(pageid, int) or pageid <= 0:
        logging.critical("fetch_wikitext_by_pageid: 'pageid' must be a positive int, got %r", pageid)
        raise SystemExit(1)

    try:
        resp = site.api(
            "query",
            pageids=str(pageid),
            prop="revisions",
            rvprop="content|timestamp",
            rvslots="main",
            formatversion="2",
            format="json",
        )
    except mwclient.errors.APIError as e:
        logging.critical("fetch_wikitext_by_pageid: API error for pageid %d: %s", pageid, e)
        raise SystemExit(1)
    except Exception as e:
        logging.critical("fetch_wikitext_by_pageid: transport/error for pageid %d: %s", pageid, e)
        raise SystemExit(1)

    # Validate structure
    query = resp.get("query")
    if not isinstance(query, dict):
        logging.critical("fetch_wikitext_by_pageid: missing 'query' in response for pageid %d: %r", pageid, resp)
        raise SystemExit(1)

    pages = query.get("pages")
    if not isinstance(pages, list) or not pages:
        logging.critical("fetch_wikitext_by_pageid: missing 'pages' array for pageid %d: %r", pageid, resp)
        raise SystemExit(1)

    page = pages[0]
    if page.get("missing"):
        logging.critical("fetch_wikitext_by_pageid: pageid %d is missing.", pageid)
        raise SystemExit(1)

    revs = page.get("revisions")
    if not isinstance(revs, list) or not revs:
        logging.critical("fetch_wikitext_by_pageid: no revisions for pageid %d.", pageid)
        raise SystemExit(1)

    main_slot = (revs[0].get("slots") or {}).get("main") or {}
    text = main_slot.get("content") or main_slot.get("*")
    ts = revs[0].get("timestamp")

    if not isinstance(text, str) or text == "":
        logging.critical("fetch_wikitext_by_pageid: empty or missing content for pageid %d.", pageid)
        raise SystemExit(1)
    if not isinstance(ts, str) or ts == "":
        logging.critical("fetch_wikitext_by_pageid: missing basetimestamp for pageid %d.", pageid)
        raise SystemExit(1)

    return {"text": text, "basetimestamp": ts}


def save_wikitext_by_pageid(site: mwclient.Site, pageid: int, new_text: str, basetimestamp: Optional[str]) -> bool:
    """
    Save transformed wikitext to a Commons File page by numeric pageid (strict mode).

    This function performs a single, guarded `action=edit` against the MediaWiki API:
      - obtains a CSRF token for the current session,
      - submits the edit by **pageid** with `assert=user`, `nocreate=1`, `bot=1`, and `maxlag=5`,
      - includes the `basetimestamp` from the *most recent read* to prevent edit conflicts.

    Strict validation:
      - `pageid` must be a positive integer.
      - `new_text` must be a non-empty string.
      - `basetimestamp` must be a non-empty ISO8601 string returned by the previous read.

    Retry policy:
      - If the API signals a token/auth issue (e.g., "badtoken", "notloggedin"), the CSRF token is
        refreshed once and the edit is retried **without** re-logging (BotPassword sessions do not
        allow mid-run logins). Any subsequent failure is treated as fatal.

    Failure behavior:
      - On configuration, validation, transport, or API errors (including unexpected response
        shapes), logs a CRITICAL message and terminates the run via SystemExit(1).

    Returns:
      - True if the edit result is "Success" (or explicitly reported as "nochange").

    Parameters:
      site (mwclient.Site): An authenticated Commons session.
      pageid (int): Numeric File page ID to edit.
      new_text (str): The complete wikitext to save.
      basetimestamp (Optional[str]): Timestamp from the latest page fetch to guard against conflicts.

    Notes:
      - Uses `assert=user` to ensure the session is authenticated (BotPassword-friendly).
      - `maxlag=5` allows the server to defer during high replica lag; callers should be prepared
        for occasional deferrals and the single token-refresh retry described above.
    """

    if not isinstance(pageid, int) or pageid <= 0:
        logging.critical("save_wikitext_by_pageid: 'pageid' must be a positive int, got %r", pageid)
        raise SystemExit(1)
    if not isinstance(new_text, str) or not new_text:
        logging.critical("save_wikitext_by_pageid: 'new_text' must be a non-empty str.")
        raise SystemExit(1)
    if not isinstance(basetimestamp, str) or not basetimestamp:
        logging.critical("save_wikitext_by_pageid: 'basetimestamp' is required and must be a non-empty str.")
        raise SystemExit(1)

    def _get_token() -> str:
        """
        Obtain a fresh CSRF (edit) token for the current MediaWiki session.

        Uses mwclient's `site.get_token('csrf')` to fetch a token suitable for
        `action=edit`. This function is intentionally strict:
          - If the token cannot be retrieved (network/API error), it logs CRITICAL
            and terminates the run (SystemExit).
          - If an empty/None token is returned, it also logs CRITICAL and terminates.

        Returns:
            str: A non-empty CSRF token string.

        Raises:
            SystemExit: If the token cannot be obtained or is empty.

        Notes:
            - CSRF tokens can expire; callers should be prepared to call this again
              and retry the edit once on token-related API errors (e.g., 'badtoken').
            - Do NOT attempt to re-login from here when using bot-password sessions;
              re-login mid-session is disallowed and should be handled at a higher level.
        """
        try:
            token = site.get_token('csrf')
        except Exception as e:
            logging.critical("save_wikitext_by_pageid: failed to obtain CSRF token: %s", e)
            raise SystemExit(1)
        if not token:
            logging.critical("save_wikitext_by_pageid: CSRF token is empty/None.")
            raise SystemExit(1)
        return token

    def _edit_with_token(token: str):
        """
        Perform a single MediaWiki 'action=edit' call using the provided CSRF token.

        This is a thin wrapper around `site.api(...)` that sets the exact parameters we want:
          - pageid (str): target page by numeric ID (avoids title encoding issues)
          - text (str): new wikitext to save
          - summary (str): EDIT_SUMMARY (global)
          - token (str): CSRF token obtained via `site.get_token('csrf')`
          - basetimestamp (str): timestamp from the last read; prevents edit conflicts
          - assert=user: ensure we are logged in (simpler and BotPassword-friendly)
          - nocreate=1: do not create a page if it doesn’t exist
          - bot=1: mark edits as bot (if the account has the right)
          - maxlag=5: be polite to the replica lag; server may defer if overloaded
          - format=json: JSON response

        Args:
            token: A non-empty CSRF token string returned by `site.get_token('csrf')`.

        Returns:
            dict: The raw JSON response from the MediaWiki API (e.g., {"edit": {"result": "Success", ...}}).

        Raises:
            mwclient.errors.APIError: If the API returns an error condition.
            Exception: For network/transport or unexpected issues.
            (These are intentionally not caught here; the caller handles retries/fail-fast.)
        """
        params = {
            "action": "edit",
            "pageid": str(pageid),
            "text": new_text,
            "summary": EDIT_SUMMARY,
            "token": token,
            "basetimestamp": basetimestamp,  # prevent edit conflicts
            # Use exactly ONE of the following assertions:
            "assert": "user",  # simpler: just assert that we're logged in
            # "assertuser": USERNAME,         # stricter: assert we are THIS user (avoid with BotPassword sessions)
            "nocreate": 1,
            "bot": 1,
            "maxlag": 5,
            "format": "json",
        }
        return site.api(**params)

    # First attempt
    token = _get_token()
    try:
        resp = _edit_with_token(token)
    except mwclient.errors.APIError as e:
        # Possible auth/token issues: assertnameduserfailed, notloggedin, badtoken, etc.
        err = str(e)
        if any(k in err for k in ("assertuserfailed", "assertnameduserfailed", "notloggedin", "badtoken")):
            logging.warning("save_wikitext_by_pageid: auth/token issue detected; re-logging and retrying once...")
            try:
                site.login(USERNAME, PASSWORD)
                token = _get_token()
                resp = _edit_with_token(token)
            except Exception as e2:
                logging.critical("save_wikitext_by_pageid: retry after re-login failed: %s", e2)
                raise SystemExit(1)
        else:
            logging.critical("save_wikitext_by_pageid: API error saving pageid %d: %s", pageid, e)
            raise SystemExit(1)
    except Exception as e:
        logging.critical("save_wikitext_by_pageid: transport/error saving pageid %d: %s", pageid, e)
        raise SystemExit(1)

    # Validate response
    if "error" in resp:
        logging.critical("save_wikitext_by_pageid: edit failed for pageid %d: %r", pageid, resp["error"])
        raise SystemExit(1)

    edit = resp.get("edit", {})
    result = edit.get("result")
    if result == "Success" or edit.get("nochange") is True:
        return True

    logging.critical("save_wikitext_by_pageid: unexpected edit response for pageid %d: %r", pageid, resp)
    raise SystemExit(1)



# ------------------------
# Pipeline
# ------------------------
def get_records() -> List[Dict]:
    """
    Build records from the Excel sheet using the Pageid column (strict mode).

    Requirements (no fallback):
      - EXCEL_FILE must exist and be readable.
      - EXCEL_SHEET must resolve to a single sheet (int index or name).
      - The DataFrame must contain EXCEL_PAGEID_COL (e.g., "Pageid").
      - Every value in EXCEL_PAGEID_COL must be a positive integer (no blanks/NaN/strings).
      - Title column (EXCEL_TITLE_COL) is optional; if missing, titles are None.

    Behavior:
      - Returns a list of dicts: {"pageid": int, "mid": "M<pageid>", "title": Optional[str]}.
      - De-duplicates by pageid while preserving first occurrence (logs a warning if duplicates were present).
      - On any validation failure, logs a CRITICAL message and aborts (SystemExit).

    Returns:
      List[Dict]: Strictly validated records ready for processing.
    """

    # Validate file exists
    if not os.path.exists(EXCEL_FILE):
        logging.critical("get_records: Excel file not found: %s", EXCEL_FILE)
        raise SystemExit(1)

    # Read sheet
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=EXCEL_SHEET)
    except Exception as e:
        logging.critical("get_records: failed to read Excel '%s' (sheet=%r): %s", EXCEL_FILE, EXCEL_SHEET, e)
        raise SystemExit(1)

    # Validate DataFrame shape
    if not hasattr(df, "columns"):
        logging.critical(
            "get_records: unexpected object returned by pandas (not a DataFrame). "
            "Check EXCEL_SHEET=%r; it must be a single sheet name or index.", EXCEL_SHEET
        )
        raise SystemExit(1)

    # Validate presence of pageid column
    if EXCEL_PAGEID_COL not in df.columns:
        logging.critical(
            "get_records: expected '%s' column not found. Columns present: %s",
            EXCEL_PAGEID_COL, list(df.columns)
        )
        raise SystemExit(1)

    # Convert pageid column to strictly positive ints (no silent coercion)
    raw_pid = df[EXCEL_PAGEID_COL]
    num_pid = pd.to_numeric(raw_pid, errors="coerce")  # floats or ints
    invalid_mask = num_pid.isna() | (num_pid <= 0)
    if invalid_mask.any():
        bad_rows = df.loc[invalid_mask, [EXCEL_PAGEID_COL]]
        # Show up to 5 problematic entries for quick debugging
        examples = bad_rows.head(5).to_dict(orient="records")
        logging.critical(
            "get_records: invalid values in '%s' (must be positive integers). "
            "Examples (first 5): %s", EXCEL_PAGEID_COL, examples
        )
        raise SystemExit(1)

    pageids = num_pid.astype(int).tolist()

    # Optional titles (for logging only)
    if EXCEL_TITLE_COL in df.columns:
        titles = df[EXCEL_TITLE_COL].astype(str).str.strip().tolist()
    else:
        titles = [None] * len(pageids)

    # Build records
    records: List[Dict] = []
    for pid, title in zip(pageids, titles):
        mid = f"M{pid}"  # mid = M + pageid
        records.append({
            "pageid": pid,              # used for API edits
            "mid": mid,                 # for logging
            "title": title or None      # for logging
        })

    # De-duplicate by pageid while preserving order
    seen = set()
    uniq: List[Dict] = []
    dup_count = 0
    for rec in records:
        pid = rec["pageid"]
        if pid in seen:
            dup_count += 1
            continue
        seen.add(pid)
        uniq.append(rec)

    if dup_count > 0:
        logging.warning("get_records: %d duplicate pageid rows were ignored (first occurrence kept).", dup_count)

    if not uniq:
        logging.critical("get_records: no valid records after validation and de-duplication.")
        raise SystemExit(1)

    # Enforce MAX_FILES strictly (do not silently process more)
    if len(uniq) > MAX_FILES:
        logging.critical("get_records: %d rows exceed MAX_FILES=%d. Reduce the input or raise MAX_FILES.",
                         len(uniq), MAX_FILES)
        raise SystemExit(1)

    return uniq

def write_status_to_excel(status_by_pageid: Dict[int, str]) -> None:
    """
    Update/append a status column in the Excel sheet:
      - Matches by EXCEL_PAGEID_COL
      - Writes statuses into EXCEL_STATUS_COL
      - Also fills/updates the ResolverURL column from Title
      - Resolves EXCEL_SHEET to an actual sheet *name* (string) before writing
      - Replaces existing sheet or creates a new one if it doesn't exist
    """


    if not status_by_pageid:
        logging.info("write_status_to_excel: nothing to write (empty status map).")
        return

    if not os.path.exists(EXCEL_FILE):
        logging.critical("write_status_to_excel: Excel file not found: %s", EXCEL_FILE)
        raise SystemExit(1)

    # --- Read the sheet as you already do
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=EXCEL_SHEET)
    except Exception as e:
        logging.critical("write_status_to_excel: failed reading Excel: %s", e)
        raise SystemExit(1)

    if EXCEL_PAGEID_COL not in df.columns:
        logging.critical("write_status_to_excel: expected '%s' column not found. Columns: %s",
                         EXCEL_PAGEID_COL, list(df.columns))
        raise SystemExit(1)

    # Ensure status column exists
    if EXCEL_STATUS_COL not in df.columns:
        df[EXCEL_STATUS_COL] = ""

    # Commented out because it only needed to run once, having completely filled the column "ResolverURL after the first run"
    # ------- Ensure & fill ResolverURL column from Title -------
    # if EXCEL_RESOLVER_COL not in df.columns:
    #     df[EXCEL_RESOLVER_COL] = ""
    #
    # filled_resolvers = 0
    # if EXCEL_TITLE_COL in df.columns:
    #     titles = df[EXCEL_TITLE_COL].astype(str).tolist()
    #     for i, t in enumerate(titles):
    #         url = resolver_from_title(t)
    #         if url:
    #             df.at[i, EXCEL_RESOLVER_COL] = url
    #             filled_resolvers += 1
    #     logging.info("write_status_to_excel: filled %d %r cells from Title.", filled_resolvers, EXCEL_RESOLVER_COL)
    # else:
    #     logging.warning("write_status_to_excel: Title column %r not present; ResolverURL not computed.", EXCEL_TITLE_COL)
    # ----------------------------------------------------------------

    # Normalize pageid column to Int64 (preserves NaN) and apply statuses
    pid_series = pd.to_numeric(df[EXCEL_PAGEID_COL], errors="coerce").astype("Int64")
    applied = 0
    for pid, status in status_by_pageid.items():
        try:
            if pid is None:
                continue
            mask = (pid_series == int(pid))
            if mask.any():
                df.loc[mask, EXCEL_STATUS_COL] = status
                applied += int(mask.sum())
        except Exception as e:
            logging.warning("write_status_to_excel: could not write status for pageid %r: %s", pid, e)

    # --- Resolve a valid *string* sheet name for writing
    try:
        wb = load_workbook(EXCEL_FILE, read_only=True)
        sheetnames = wb.sheetnames
        wb.close()
    except Exception as e:
        logging.critical("write_status_to_excel: cannot read workbook sheet names: %s", e)
        raise SystemExit(1)

    # Determine target sheet name (string)
    if isinstance(EXCEL_SHEET, int):
        if EXCEL_SHEET < 0 or EXCEL_SHEET >= len(sheetnames):
            logging.critical("write_status_to_excel: sheet index %r out of range. Available sheets: %s",
                             EXCEL_SHEET, sheetnames)
            raise SystemExit(1)
        target_sheet = sheetnames[EXCEL_SHEET]
    elif isinstance(EXCEL_SHEET, str) and EXCEL_SHEET.strip():
        target_sheet = EXCEL_SHEET
    else:
        # default to the first sheet's actual name
        target_sheet = sheetnames[0]

    # Choose replace vs new
    sheet_exists = target_sheet in sheetnames
    mode = "a"
    if_sheet_exists = "replace" if sheet_exists else "new"

    # --- Write back
    try:
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode=mode, if_sheet_exists=if_sheet_exists) as writer:
            df.to_excel(writer, sheet_name=target_sheet, index=False)
        logging.info("write_status_to_excel: wrote %d status cells to '%s' (sheet=%r).",
                     applied, EXCEL_FILE, target_sheet)
    except Exception as e:
        logging.critical("write_status_to_excel: failed writing Excel: %s", e)
        raise SystemExit(1)

def open_page_in_browser(pageid: int, title: Optional[str] = None) -> None:
    """
    Open the Commons File page in the system's default web browser by *pageid*.

    This uses the stable redirect endpoint:
        https://commons.wikimedia.org/wiki/Special:Redirect/page/{pageid}
    which avoids any title-encoding issues and remains valid even if the file title
    changes (move/rename), because it targets the page *by id*.

    Args:
        pageid (int):
            Numeric MediaWiki page ID (namespace 6 / File:). Must be a positive integer.
        title (Optional[str]):
            Optional human-readable label used only for logging (has no effect on the URL).

    Behavior:
        - Logs the URL being opened (and the Title if provided).
        - Attempts to open a new browser tab via Python's `webbrowser` module.
        - Any exception from the browser invocation is caught and logged as a warning;
          the function does not raise.

    Returns:
        None

    Notes:
        - This function is typically called after a successful edit to let you
          visually verify the change. Whether it is called is controlled elsewhere
          (e.g., OPEN_AFTER_SUCCESS / OPEN_AFTER_SUCCESS_MAX).
        - `webbrowser.open_new_tab()` is best-effort and platform-dependent; it may
          return without actually focusing the browser, especially in headless or
          restricted environments.
        - Using Special:Redirect by pageid avoids URL encoding of file titles and is
          robust against future title changes.

    Example:
        open_page_in_browser(109018562, "Nieuw_Amsterdamsch_handels-_en_effectenblad_07-11-1859_(IA_...).pdf")
    """
    try:
        url = f"https://commons.wikimedia.org/wiki/Special:Redirect/page/{pageid}"
        logging.info("Opening in browser: %s%s",
                     url, f" (Title={title!r})" if title else "")
        webbrowser.open_new_tab(url)
    except Exception as e:
        logging.warning("Failed to open browser for pageid %d: %s", pageid, e)


def process_records(site: mwclient.Site, records: List[Dict]) -> Dict[int, str]:
    """
    Process each record and return a status map {pageid: status}.

    Status values (globals expected):
      - STATUS_PROCESSED
      - STATUS_SKIPPED_NO_OLD
      - STATUS_ERROR

    Behavior:
      - Logs a clear separator line (50 '=') between files.
      - Opens the Commons page in a browser tab after a successful edit if OPEN_AFTER_SUCCESS is truthy,
        up to OPEN_AFTER_SUCCESS_MAX tabs per run (both read from globals).
      - Sleeps EDIT_SLEEP_SEC seconds between files if configured (global).
      - After every CHECKPOINT_EVERY_SUCCESS *successful edits* (default=50), writes accumulated statuses
        to Excel via write_status_to_excel() (checkpointing).
      - Continues across errors, marking STATUS_ERROR for that page.
    """
    if not isinstance(records, list):
        logging.critical("process_records: 'records' must be a list, got %r", type(records).__name__)
        raise SystemExit(1)
    if len(records) == 0:
        logging.critical("process_records: no records to process.")
        raise SystemExit(1)

    total = len(records)
    sep = "=" * 50
    status_by_pageid: Dict[int, str] = {}

    # Open-after-success toggles (read safely from globals)
    open_after = bool(globals().get("OPEN_AFTER_SUCCESS", 1))
    open_after_max = int(globals().get("OPEN_AFTER_SUCCESS_MAX", 10))
    opened = 0

    # Optional polite delay
    sleep_sec = globals().get("EDIT_SLEEP_SEC", 0)

    # NEW: checkpointing config
    checkpoint_every = int(globals().get("CHECKPOINT_EVERY_SUCCESS", 50))
    successes_since_flush = 0
    total_successes = 0

    for idx, rec in enumerate(tqdm(records, desc="Processing files", unit="file"), start=1):
        # Validate record shape
        if not isinstance(rec, dict) or "pageid" not in rec:
            logging.critical("process_records: invalid record at index %d: %r", idx, rec)
            status_by_pageid[-1] = STATUS_ERROR  # placeholder so caller sees something went wrong
            raise SystemExit(1)

        pid = rec["pageid"]
        if not isinstance(pid, int) or pid <= 0:
            logging.critical("process_records: invalid 'pageid' at index %d: %r", idx, pid)
            status_by_pageid[-1] = STATUS_ERROR
            raise SystemExit(1)

        title = rec.get("title")
        mid   = rec.get("mid") or f"M{pid}"
        label = title if (isinstance(title, str) and title.strip()) else f"pageid {pid}"

        logging.info("(%d/%d) Processing: Title=%r, MID=%s, Pageid=%d", idx, total, title, mid, pid)

        try:
            fetched = fetch_wikitext_by_pageid(site, pid)
            text = fetched["text"]
            base = fetched["basetimestamp"]

            new_text = transform_wikitext(text)
            if new_text == text:
                logging.info("%d - No change needed (no OLD pattern): %s", idx, label)
                status_by_pageid[pid] = STATUS_SKIPPED_NO_OLD
            else:
                save_wikitext_by_pageid(site, pid, new_text, base)
                logging.info("%d - Updated: %s", idx, label)
                status_by_pageid[pid] = STATUS_PROCESSED

                # NEW: count successes and checkpoint if threshold reached
                total_successes += 1
                successes_since_flush += 1
                if checkpoint_every > 0 and successes_since_flush >= checkpoint_every:
                    try:
                        write_status_to_excel(status_by_pageid)
                        logging.info("Checkpoint: wrote statuses after %d successful edits (total successes so far: %d).",
                                     successes_since_flush, total_successes)
                    except SystemExit:
                        raise
                    except Exception as e:
                        logging.warning("Checkpoint write failed (continuing): %s", e)
                    finally:
                        successes_since_flush = 0  # reset counter regardless of write outcome

                # Open browser tab after success (respect caps/toggles)
                if open_after and opened < open_after_max:
                    opener = globals().get("open_page_in_browser")
                    if callable(opener):
                        opener(pid, title)
                        opened += 1
                    else:
                        logging.debug("open_page_in_browser not available; skipping browser open.")

        except SystemExit as se:
            logging.error("%d - Error processing %s (continuing): %s", idx, label, se)
            status_by_pageid[pid] = STATUS_ERROR
        except Exception as e:
            logging.error("%d - Unexpected error processing %s (continuing): %s", idx, label, e)
            status_by_pageid[pid] = STATUS_ERROR
        finally:
            if idx < total:
                logging.info(sep)
            # Optional polite sleep
            try:
                if sleep_sec and float(sleep_sec) > 0 and idx < total:
                    time.sleep(float(sleep_sec))
            except Exception:
                pass

    # NEW: flush any remaining successes that didn’t hit the threshold
    if checkpoint_every > 0 and successes_since_flush > 0:
        try:
            write_status_to_excel(status_by_pageid)
            logging.info("Final checkpoint: wrote statuses after last %d successful edits (total successes: %d).",
                         successes_since_flush, total_successes)
        except SystemExit:
            raise
        except Exception as e:
            logging.warning("Final checkpoint write failed (continuing): %s", e)

    return status_by_pageid


def main() -> None:
    """
    Orchestrate a full, strict update run driven by the Excel input.

    High-level flow:
      1) Read & validate records from the Excel via get_records() (strict):
         - Requires EXCEL_FILE/EXCEL_SHEET and EXCEL_PAGEID_COL to be valid.
         - Produces a list of {"pageid", "mid", "title"} dicts.
      2) Optionally slice workload via apply_slice_records() using HEAD or RANGE (strict, mutually exclusive).
      3) Log in to Wikimedia Commons using mwclient with USERNAME/PASSWORD and USER_AGENT.
      4) Log the selected workload (Title, MID, Pageid) and print a separator line.
      5) Process each record via process_records():
         - Fetch wikitext by pageid, transform exactly one OLD source block, save the page.
         - Optionally open each successfully edited page in a browser tab (OPEN_AFTER_SUCCESS / OPEN_AFTER_SUCCESS_MAX).
         - Optionally sleep between files (EDIT_SLEEP_SEC) and pass maxlag=5 to be polite to the API.
         - Returns a per-page status map (processed / skipped_no_old_pattern / skipped_error).
      6) Write statuses back into the Excel sheet via write_status_to_excel() (creates/updates EXCEL_STATUS_COL).

    Strictness & failure behavior:
      - Most validation and API anomalies raise SystemExit with a clear CRITICAL log message.
      - process_records() continues across per-file errors so that statuses can be written.
      - This function re-raises SystemExit to preserve strict abort semantics; unexpected exceptions are logged
        and converted to SystemExit(1).

    Side effects:
      - Network access to the Commons API for read/edit operations.
      - Writes an updated sheet back to EXCEL_FILE (replacing or creating the target sheet).
      - May open browser tabs (platform-dependent) if OPEN_AFTER_SUCCESS is enabled.
      - May sleep briefly between edits (EDIT_SLEEP_SEC).

    Environment variables used (see module docstring for defaults):
      - WIKIMEDIA_USERNAME, WIKIMEDIA_PASSWORD, WIKIMEDIA_USER_AGENT
      - EXCEL_FILE, EXCEL_SHEET, EXCEL_PAGEID_COL, EXCEL_STATUS_COL
      - MAX_FILES, EDIT_SUMMARY, EDIT_SLEEP_SEC, HEAD, RANGE
      - OPEN_AFTER_SUCCESS, OPEN_AFTER_SUCCESS_MAX

    Returns:
      None

    Raises:
      SystemExit: on fatal configuration/API errors (see strictness notes above).
    """
    try:
        records = get_records()
        records = apply_slice_records(records)
        if not records:
            logging.critical("main: no records to process after reading/slicing.")
            raise SystemExit(1)

        # Login
        try:
            site = mwclient.Site("commons.wikimedia.org", clients_useragent=USER_AGENT)
            site.login(USERNAME, PASSWORD)
        except mwclient.LoginError as e:
            logging.critical("main: login failed: %s", e)
            raise SystemExit(1)
        except Exception as e:
            logging.critical("main: unexpected error during login: %s", e)
            raise SystemExit(1)

        logging.info("Selected %d files to process.", len(records))
        for rec in records:
            logging.info(
                " - Will process: Title=%r, MID=%s, Pageid=%d",
                rec.get("title"), rec.get("mid"), rec.get("pageid")
            )
        logging.info("=" * 50)

        # Process and collect statuses
        status_by_pageid = process_records(site, records)

        # Write statuses back to Excel
        write_status_to_excel(status_by_pageid)

        logging.info("Finished processing.")

    except SystemExit:
        raise
    except Exception as e:
        logging.critical("main: unexpected top-level error: %s", e)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
